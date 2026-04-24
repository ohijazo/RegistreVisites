import io
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.db.models import Visit

EXPORT_COLUMNS = [
    ("ID", lambda v: str(v.id)),
    ("Nom", lambda v: v.first_name),
    ("Cognoms", lambda v: v.last_name),
    ("Empresa", lambda v: v.company),
    ("Telèfon", lambda v: v.phone or ""),
    ("Departament", lambda v: getattr(v.department, "name_ca", "") if v.department else ""),
    ("Motiu visita", lambda v: v.visit_reason),
    ("Idioma", lambda v: v.language),
    ("Data entrada", lambda v: v.checked_in_at.strftime("%d/%m/%Y %H:%M") if v.checked_in_at else ""),
    ("Data sortida", lambda v: v.checked_out_at.strftime("%d/%m/%Y %H:%M") if v.checked_out_at else ""),
    ("Durada (min)", lambda v: round((v.checked_out_at - v.checked_in_at).total_seconds() / 60) if v.checked_out_at and v.checked_in_at else ""),
    ("Mètode sortida", lambda v: v.checkout_method or ""),
    ("RGPD acceptat", lambda v: v.accepted_at.strftime("%d/%m/%Y %H:%M") if v.accepted_at else ""),
]


def visits_to_excel(visits: list[Visit], filename_date_range: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Visites"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    headers = [col[0] for col in EXPORT_COLUMNS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, visit in enumerate(visits, 2):
        for col_idx, (_, extractor) in enumerate(EXPORT_COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=extractor(visit))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def visits_to_csv(visits: list[Visit]) -> io.BytesIO:
    buffer = io.BytesIO()
    wrapper = io.TextIOWrapper(buffer, encoding="utf-8-sig", newline="")
    writer = csv.writer(wrapper, delimiter=";")

    writer.writerow([col[0] for col in EXPORT_COLUMNS])
    for visit in visits:
        writer.writerow([extractor(visit) for _, extractor in EXPORT_COLUMNS])

    wrapper.flush()
    wrapper.detach()
    buffer.seek(0)
    return buffer
